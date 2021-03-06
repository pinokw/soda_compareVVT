from openpyxl import load_workbook

def read_sheet(sheet, table_data):
    header=[]
    a=0
    for row1 in sheet.iter_rows(max_row=1):
        for cell in row1:
            header.append(cell.value)
    for row2 in sheet.iter_rows(min_row=2, values_only=True):
        table_entry={}
        for x, cell in enumerate(row2):       
            info=cell
            table_entry[header[x]]=info
        table_data.append(table_entry)
        a=a+1
    return table_data

def readxls_main(xlsfile):
    wb=load_workbook(xlsfile)
    names_sht=wb.sheetnames
    xls_data={}
    for sh_name in names_sht:
        table_data=[]
        sheet=wb[sh_name]
        tableentry=read_sheet(sheet, table_data)
        xls_data[sh_name]=tableentry
    return xls_data


if (__name__=='__main__'):
    #readxls_main(r"C:\Users\DrAndreasPinheiro\Documents\Code\soda_compareVVT\SODA_Vorlage_Allgemein_v6.xlsx")
    pass