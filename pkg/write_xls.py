from openpyxl import Workbook
import os, sys
from pathlib import Path
fileDir = os.path.dirname(os.path.realpath('__file__'))
sys.path.append(fileDir)


def print_xls(vvt_check, tom_check, todo_check, str_datei):
    if str_datei == "":
        str_xlsfile=os.path.join(fileDir, 'compare_result.xlsx')
        file_xls=Path(str_xlsfile)
        if file_xls.is_file():
            os.remove(str_xlsfile)
    else:
        str_datei=str_datei + '.xlsx'
        str_xlsfile=os.path.join(fileDir, str_datei)
    wb=Workbook()
    #Sheet VVT
    sheet = wb.create_sheet('VVT', 0)
    headerdata_vvt=('Bezeichnung', 'Löschfrist', 'Risiko für Betroffenen',\
        'Beschreibung des Vorgangs und Zweck der Verarbeitung', 'Rechtsgrundlagen',\
        'Rechtliche Würdigung', 'Inhalte ALT', 'Inhalte NEU')
    sheet.append(headerdata_vvt)
    for a, vvt_entry in enumerate(vvt_check):
        sheet.cell(row=a+2, column=1, value=vvt_entry['Bezeichnung'])
        if 'Löschfrist' in vvt_entry:
            sheet.cell(row=a+2, column=2, value=vvt_entry['Löschfrist'])
        if 'Risiko für Betroffenen' in vvt_entry:
            sheet.cell(row=a+2, column=3, value=vvt_entry['Risiko für Betroffenen'])
        if 'Beschreibung des Vorgangs und Zweck der Verarbeitung' in vvt_entry:
            sheet.cell(row=a+2, column=4, value=vvt_entry['Beschreibung des Vorgangs und Zweck der Verarbeitung'])
        if 'Rechtsgrundlagen' in vvt_entry:
            sheet.cell(row=a+2, column=5, value=vvt_entry['Rechtsgrundlagen'])
        if 'Rechtliche Würdigung' in vvt_entry:
            sheet.cell(row=a+2, column=6, value=vvt_entry['Rechtliche Würdigung'])
        if 'Text_OLD' in vvt_entry:
            sheet.cell(row=a+2, column=7, value=vvt_entry['Text_OLD'])
        if 'Text_NEW' in vvt_entry:
            sheet.cell(row=a+2, column=8, value=vvt_entry['Text_NEW'])
    freezy=sheet['Y2']
    sheet.freeze_panes=freezy


    #Sheet TOM
    sheet1 = wb.create_sheet('TOM', 1)
    headerdata_tom=('Allgemeine TOM Nr', 'Kategorie der allg. TOM', 'Beschreibung')
    sheet1.append(headerdata_tom)
    for b, tom_entry in enumerate(tom_check):
        sheet1.cell(row=b+2, column=1, value=tom_entry['Allgemeine TOM Nr'])
        sheet1.cell(row=b+2, column=2, value=tom_entry['Kategorie der allg. TOM'])
        sheet1.cell(row=b+2, column=3, value=tom_entry['Beschreibung'])
    freezy=sheet1['Y2']
    sheet1.freeze_panes=freezy


    #Sheet ToDo
    sheet2 = wb.create_sheet('ToDos', 2)
    headerdata_todo=('Verfahren Nr', 'Neue Bezeichnung', 'Neue Beschreibung', 'Status')
    sheet2.append(headerdata_todo)
    for c, todo_entry in enumerate(todo_check):
        sheet2.cell(row=c+2, column=1, value=todo_entry['Verfahren Nr'])
        sheet2.cell(row=c+2, column=2, value=todo_entry['Bezeichnung'])
        sheet2.cell(row=c+2, column=3, value=todo_entry['Beschreibung'])
        sheet2.cell(row=c+2, column=4, value=todo_entry['Erledigt'])
    freezy=sheet2['Y2']
    sheet2.freeze_panes=freezy
    #sheet.auto_filter.ref = "A1:Y1"
    wb.save(str_xlsfile)

def find_column_vvt(vvt_entry):
    pass
